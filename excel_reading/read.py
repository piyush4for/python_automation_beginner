import openpyxl, os
os.chdir('F:\course\python autmation\excel_reading')

#reading excel file
workbook = openpyxl.load_workbook('example.xlsx')
print(workbook.sheetnames)
sheet = workbook['Sheet1']

print(str(sheet['B1'].value))
print(str(sheet.cell(row=1,column=2).value))

for i in range(1,8):
    print(i, sheet.cell(row=i,column=2).value)
