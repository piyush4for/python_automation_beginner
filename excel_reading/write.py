import openpyxl,os
wb = openpyxl.Workbook()

print(wb.sheetnames)
sheet = wb['Sheet']

sheet['A1'] = 32
sheet['A2'] = 'hello'
wb.save('example2.xlsx')

sheet2= wb.create_sheet()
print(wb.sheetnames)